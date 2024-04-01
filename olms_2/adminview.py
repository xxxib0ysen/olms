from PyQt5.QtWidgets import *
from admin.view.student_right import StudentRight
from admin.view.teacher_right import TeacherRight
from admin.view.face_right import FaceRight
from server import *
import config as c
import openpyxl
import requests

ADMIN_WIDTH = 1000
ADMIN_HEIGHT = 660


class AdminView(QWidget):
    def __init__(self, access_token, return_login):
        super().__init__()
        self.access_token = access_token
        self.return_login = return_login
        self.setWindowTitle("管理员界面")
        self.setFixedSize(ADMIN_WIDTH, ADMIN_HEIGHT)
        # 获取屏幕大小, 设置居中
        center = QDesktopWidget().availableGeometry().center()
        self.move(center.x() - ADMIN_WIDTH / 2, center.y() - ADMIN_HEIGHT / 2)

        self.left_frame = QFrame(self)
        self.left_frame.setGeometry(20, 30, 180, 600)

        btn_txt = ('学生管理', '教师管理', '人脸管理', '注销')
        btn_command = (
            self.on_student,
            self.on_teacher,
            self.on_face,
            self.return_login
        )
        for i in range(len(btn_txt)):
            btn = QPushButton(btn_txt[i], self.left_frame)
            btn.setGeometry(18, i * 70 + 10, 130, 40)
            btn.clicked.connect(btn_command[i])

        self.student_left = QFrame(self)
        self.student_left.setGeometry(20, 30, 180, 600)
        self.student_left.hide()
        btn_txt = ('新增', '编辑', '删除', '导出学生信息', '导入学生信息', '返回')
        btn_command = (
            self.on_student_add, self.on_student_edit, self.on_student_del,
            self.on_student_export, self.on_student_import, self.on_return,
        )
        for i in range(len(btn_txt)):
            btn = QPushButton(btn_txt[i], self.student_left)
            btn.setGeometry(18, i * 70 + 10, 130, 40)
            btn.clicked.connect(btn_command[i])
        self.student_view = StudentRight('操作区域', self)
        self.student_view.hide()

        self.teacher_left = QFrame(self)
        self.teacher_left.setGeometry(20, 30, 180, 600)
        self.teacher_left.hide()
        btn_txt = ('新增', '编辑', '删除', '导出教师信息', '导入教师信息', '返回')
        btn_command = (
            self.on_teacher_add, self.on_teacher_edit, self.on_teacher_del,
            self.on_teacher_export, self.on_teacher_import, self.on_return,
        )
        for i in range(len(btn_txt)):
            btn = QPushButton(btn_txt[i], self.teacher_left)
            btn.setGeometry(18, i * 70 + 10, 130, 40)
            btn.clicked.connect(btn_command[i])
        self.teacher_view = TeacherRight('操作区域', self)
        self.teacher_view.hide()

        self.face_left = QFrame(self)
        self.face_left.setGeometry(20, 30, 180, 600)
        self.face_left.hide()
        btn_txt = ('删除信息', '返回')
        btn_command = (self.on_face_del, self.on_return)
        for i in range(len(btn_txt)):
            btn = QPushButton(btn_txt[i], self.face_left)
            btn.setGeometry(18, i * 70 + 10, 130, 40)
            btn.clicked.connect(btn_command[i])
        self.face_view = FaceRight('操作区域', self)
        self.face_view.hide()

    # 删除选中人脸
    def on_face_del(self):
        user, face_token = self.face_view.get_current_select()
        if user is None or face_token is None:
            return
        request_url = c.Delete_face_url
        params = {
            "user_id": user,
            "group_id": "aaa",
            "face_token": face_token
        }
        request_url = request_url + "?access_token=" + self.access_token
        headers = {'content-type': 'application/json'}
        response = requests.post(request_url, data=params, headers=headers)
        if response:
            data = response.json()
            if data['error_code'] == 0:
                QMessageBox.about(self, "人脸删除结果", "人脸删除成功")
                self.face_view.update_data(self.access_token)
            else:
                QMessageBox.about(self, "人脸删除结果", "人脸删除失败" + data['error_msg'])

    # 新增学生
    def on_student_add(self):
        add_student_dialog = AddStudentDialog()
        if add_student_dialog.exec_() == QDialog.Accepted:
            # 刷新
            self.load_student_data()

    # 编辑学生
    def on_student_edit(self):
        try:
            student_info = self.student_view.get_row_data()
            if student_info is not None:
                # 显示编辑学生信息的对话框
                edit_student_dialog = EditStudentDialog(student_info)
                if edit_student_dialog.exec_() == QDialog.Accepted:
                    self.load_student_data()
            else:
                QMessageBox.warning(self, "警告", "请选择要编辑的学生。")
        except Exception as e:
            print("An error occurred:", str(e))

    # 删除学生
    def on_student_del(self):
        try:
            student_info = self.student_view.get_row_data()
            if student_info is not None:
                sid = student_info['sid']
                confirm = QMessageBox.question(self, "确认删除", "是否确定删除选定的学生记录？", QMessageBox.Yes | QMessageBox.No)
                if confirm == QMessageBox.Yes:
                    delete_student(sid)
                    self.load_student_data()
            else:
                QMessageBox.warning(self, "警告", "请选择要删除的学生。")
        except Exception as e:
            print("An error occurred:", str(e))

    # 导出学生数据
    def on_student_export(self):
        def export_students_to_excel(file_name):
            # 导出学生信息到Excel文件
            try:
                students = get_all_students()
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["学号", "姓名", "密码", "班级", "课程", "总分"])
                for student in students:
                    sheet.append(student)
                workbook.save(file_name)
                return True
            except Exception as e:
                print("Error:", e)
                return False

        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "导出学生信息", "", "Excel Files (*.xlsx);;All Files (*)",
                                                   options=options)
        if file_name:
            success = export_students_to_excel(file_name)
            if success:
                QMessageBox.information(self, "导出成功", "学生信息已成功导出。")
            else:
                QMessageBox.warning(self, "导出失败", "导出学生信息时出现错误。")

    # 导入学生数据
    def on_student_import(self):
        def import_students_from_excel(file_name):
            try:
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    sid, sname, password, sclass, cname, overall_score = row
                    student = Student(sid, sname, password, sclass, cname, overall_score)
                    add_student(student)
                return True
            except Exception as e:
                print("Error:", e)
                return False

        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "导入学生信息", "", "Excel Files (*.xlsx);;All Files (*)",
                                                   options=options)
        if file_name:
            success = import_students_from_excel(file_name)
            if success:
                QMessageBox.information(self, "导入成功", "学生信息已成功导入。")
                self.load_student_data()
            else:
                QMessageBox.warning(self, "导入失败", "导入学生信息时出现错误。")

    def load_student_data(self):
        students = get_all_students_by_sid()
        self.student_view.update_data(students)

    def on_teacher_add(self):
        add_teacher_dialog = AddTeacherDialog()
        if add_teacher_dialog.exec_() == QDialog.Accepted:
            self.load_teacher_data()

    def on_teacher_edit(self):
        try:
            teacher_info = self.teacher_view.get_row_data()
            if teacher_info is not None:
                # 显示编辑学生信息的对话框
                edit_teacher_dialog = EditTeacherDialog(teacher_info)
                if edit_teacher_dialog.exec_() == QDialog.Accepted:
                    self.load_teacher_data()
            else:
                QMessageBox.warning(self, "警告", "请选择要编辑的教师。")
        except Exception as e:
            print("An error occurred:", str(e))

    def on_teacher_del(self):
        try:
            teacher_info = self.teacher_view.get_row_data()
            if teacher_info is not None:
                tid = teacher_info['tid']
                confirm = QMessageBox.question(self, "确认删除", "是否确定删除选定的教师记录？", QMessageBox.Yes | QMessageBox.No)
                if confirm == QMessageBox.Yes:
                    delete_teacher(tid)  # 调用管理函数删除教师
                    self.load_teacher_data()
            else:
                QMessageBox.warning(self, "警告", "请选择要删除的教师。")
        except Exception as e:
            print("An error occurred:", str(e))

    def on_teacher_export(self):
        def export_teachers_to_excel(file_name):
            # 导出教师信息到Excel文件
            try:
                teachers = get_all_teachers()  # 获取所有教师信息
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["教工号", "姓名", "密码", "课程"])
                for teacher in teachers:
                    sheet.append(teacher)
                workbook.save(file_name)
                return True
            except Exception as e:
                print("Error:", e)
                return False

        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "导出教师信息", "", "Excel Files (*.xlsx);;All Files (*)",
                                                   options=options)
        if file_name:
            success = export_teachers_to_excel(file_name)
            if success:
                QMessageBox.information(self, "导出成功", "教师信息已成功导出。")
            else:
                QMessageBox.warning(self, "导出失败", "导出教师信息时出现错误。")

    def on_teacher_import(self):
        def import_teachers_from_excel(file_name):
            try:
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    tid, tname, password, cname = row
                    teacher = Teacher(tid, tname, password, cname)
                    add_teacher(teacher)
                return True
            except Exception as e:
                print("Error:", e)
                return False

        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "导入教师信息", "", "Excel Files (*.xlsx);;All Files (*)",
                                                   options=options)
        if file_name:
            success = import_teachers_from_excel(file_name)
            if success:
                QMessageBox.information(self, "导入成功", "教师信息已成功导入。")
                self.load_teacher_data()
            else:
                QMessageBox.warning(self, "导入失败", "导入教师信息时出现错误。")

    def load_teacher_data(self):
        teachers = get_all_teachers()
        self.teacher_view.update_data(teachers)

    def load_face_data(self):
        self.face_view.update_data(self.access_token)


    def on_student(self):
        self.left_frame.hide()
        # 隐藏教师的界面
        self.teacher_left.hide()
        self.teacher_view.hide()
        # 隐藏人脸的界面
        self.face_left.hide()
        self.face_view.hide()
        # 显示学生的界面
        self.student_left.show()
        self.student_view.show()
        # 更新数据
        self.load_student_data()

    def on_teacher(self):
        self.left_frame.hide()
        # 隐藏学生的界面
        self.student_left.hide()
        self.student_view.hide()
        # 隐藏人脸的界面
        self.face_left.hide()
        self.face_view.hide()
        # 显示教师的界面
        self.teacher_left.show()
        self.teacher_view.show()
        # 更新数据
        self.load_teacher_data()

    def on_face(self):
        self.left_frame.hide()
        # 隐藏学生的界面
        self.student_left.hide()
        self.student_view.hide()
        # 显示人脸的界面
        self.face_left.show()
        self.face_view.show()
        # 隐藏教师的界面
        self.teacher_left.hide()
        self.teacher_view.hide()
        # 更新数据
        self.load_face_data()

    def on_return(self):
        self.left_frame.show()
        self.student_left.hide()
        self.student_view.hide()
        self.teacher_left.hide()
        self.teacher_view.hide()
        self.face_left.hide()
        self.face_view.hide()

    def logout(self):
        # 关闭当前窗口
        self.close()


class AddStudentDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("添加学生")
        self.setFixedSize(300, 350)
        center = QDesktopWidget().availableGeometry().center()
        self.move(center.x() - 300 / 2, center.y() - 300 / 2)
        layout = QVBoxLayout()
        sid_label = QLabel("学号:")
        self.sid_input = QLineEdit()
        sname_label = QLabel("姓名:")
        self.sname_input = QLineEdit()
        password_label = QLabel("密码:")
        self.password_input = QLineEdit()
        sclass_label = QLabel("班级:")
        self.sclass_input = QLineEdit()
        cname_label = QLabel("课程:")
        self.cname_input = QLineEdit()
        score_label = QLabel("总分:")
        self.score_input = QLineEdit()
        add_button = QPushButton("添加")
        add_button.clicked.connect(self.add_student)
        layout.addWidget(sid_label)
        layout.addWidget(self.sid_input)
        layout.addWidget(sname_label)
        layout.addWidget(self.sname_input)
        layout.addWidget(password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(sclass_label)
        layout.addWidget(self.sclass_input)
        layout.addWidget(cname_label)
        layout.addWidget(self.cname_input)
        layout.addWidget(score_label)
        layout.addWidget(self.score_input)
        layout.addWidget(add_button)
        self.setLayout(layout)

    def add_student(self):
        sid = self.sid_input.text()
        sname = self.sname_input.text()
        password = self.password_input.text()
        sclass = self.sclass_input.text()
        cname = self.cname_input.text()
        overall_score = self.score_input.text()
        new_student = Student(sid, sname, password, sclass, cname, overall_score)
        add_student(new_student)
        self.accept()


class EditStudentDialog(QDialog):
    def __init__(self, student_info):
        super().__init__()
        self.setWindowTitle("编辑学生信息")
        self.setFixedSize(400, 450)
        center = QDesktopWidget().availableGeometry().center()
        self.move(center.x() - 300 / 2, center.y() - 300 / 2)

        self.student_info = student_info
        layout = QVBoxLayout()

        sid_label = QLabel("学号:")
        self.sid_input = QLineEdit(self)
        self.sid_input.setText(student_info["sid"])
        layout.addWidget(sid_label)
        layout.addWidget(self.sid_input)

        student_name_label = QLabel("姓名:")
        self.student_name_input = QLineEdit(self)
        self.student_name_input.setText(student_info["sname"])
        layout.addWidget(student_name_label)
        layout.addWidget(self.student_name_input)

        password_label = QLabel("密码:")
        self.password_input = QLineEdit(self)
        self.password_input.setText(student_info["password"])
        layout.addWidget(password_label)
        layout.addWidget(self.password_input)

        sclass_label = QLabel("班级:")
        self.sclass_input = QLineEdit(self)
        self.sclass_input.setText(student_info["sclass"])
        layout.addWidget(sclass_label)
        layout.addWidget(self.sclass_input)

        cname_label = QLabel("课程:")
        self.cname_input = QLineEdit(self)
        self.cname_input.setText(student_info["cname"])
        layout.addWidget(cname_label)
        layout.addWidget(self.cname_input)

        score_label = QLabel("总分:")
        self.score_input = QLineEdit(self)
        self.score_input.setText(student_info["overall_score"])
        layout.addWidget(score_label)
        layout.addWidget(self.score_input)

        save_button = QPushButton("保存", self)
        save_button.clicked.connect(self.save_student_info)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def save_student_info(self):
        try:
            edited_student_info = {
                "sid": self.sid_input.text(),
                "sname": self.student_name_input.text(),
                "password": self.password_input.text(),
                "sclass": self.sclass_input.text(),
                "cname": self.cname_input.text(),
                "overall_score": self.score_input.text()
            }

            edit_student(self.student_info["sid"], edited_student_info)
            self.accept()
        except Exception as e:
            print("An error occurred:", str(e))


class AddTeacherDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("添加教师")
        self.setFixedSize(300, 250)

        layout = QVBoxLayout()
        tid_label = QLabel("教工号:")
        self.tid_input = QLineEdit()
        tname_label = QLabel("姓名:")
        self.tname_input = QLineEdit()
        password_label = QLabel("密码:")
        self.password_input = QLineEdit()
        cname_label = QLabel("课程:")
        self.cname_input = QLineEdit()
        add_button = QPushButton("添加")
        add_button.clicked.connect(self.add_teacher)

        layout.addWidget(tid_label)
        layout.addWidget(self.tid_input)
        layout.addWidget(tname_label)
        layout.addWidget(self.tname_input)
        layout.addWidget(password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(cname_label)
        layout.addWidget(self.cname_input)
        layout.addWidget(add_button)
        self.setLayout(layout)

    def add_teacher(self):
        tid = self.tid_input.text()
        tname = self.tname_input.text()
        password = self.password_input.text()
        cname = self.cname_input.text()
        new_teacher = Teacher(tid, tname, password, cname)
        add_teacher(new_teacher)
        self.accept()


class EditTeacherDialog(QDialog):
    def __init__(self, teacher_info):
        super().__init__()
        self.setWindowTitle("编辑教师信息")
        self.setFixedSize(300, 250)

        layout = QVBoxLayout()
        tid_label = QLabel("教工号:")
        self.tid_input = QLineEdit(self)
        self.tid_input.setText(teacher_info["tid"])

        tname_label = QLabel("姓名:")
        self.tname_input = QLineEdit(self)
        self.tname_input.setText(teacher_info["tname"])

        password_label = QLabel("密码:")
        self.password_input = QLineEdit(self)
        self.password_input.setText(teacher_info["password"])

        cname_label = QLabel("课程:")
        self.cname_input = QLineEdit(self)
        self.cname_input.setText(teacher_info["cname"])

        save_button = QPushButton("保存", self)
        save_button.clicked.connect(self.save_teacher_info)

        layout.addWidget(tid_label)
        layout.addWidget(self.tid_input)
        layout.addWidget(tname_label)
        layout.addWidget(self.tname_input)
        layout.addWidget(password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(cname_label)
        layout.addWidget(self.cname_input)
        layout.addWidget(save_button)
        self.setLayout(layout)

    def save_teacher_info(self):
        try:
            edited_teacher_info = {
                "tid": self.tid_input.text(),
                "tname": self.tname_input.text(),
                'password': self.password_input.text(),
                "cname": self.cname_input.text()
            }
            edit_teacher(self.teacher_info["tid"], edited_teacher_info)  # 调用编辑教师信息的管理函数
            self.accept()
        except Exception as e:
            print("An error occurred:", str(e))


# student
class Student:
    def __init__(self, sid, sname, password, sclass, cname, overall_score):
        self.sid = sid
        self.sname = sname
        self.password = password
        self.sclass = sclass
        self.cname = cname
        self.overall_score = overall_score


class Teacher:
    def __init__(self, tid, tname, password, cname):
        self.tid = tid
        self.tname = tname
        self.password = password
        self.cname = cname
